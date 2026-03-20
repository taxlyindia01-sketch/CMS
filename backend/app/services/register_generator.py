"""
Statutory Register Generator — Module 4
Generates all 4 statutory registers as PDF and Excel:
  1. Register of Members (MGT-1)
  2. Register of Directors & KMP (MBP-1 / DIR-12)
  3. Register of Charges (CHG-7)
  4. Register of Share Transfers (SH-6)

Uses ReportLab for PDF and openpyxl for Excel.
"""

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

# ─── ReportLab imports ────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ─── openpyxl imports ─────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════════════

NAVY    = colors.HexColor("#0f1f3d")
GOLD    = colors.HexColor("#e8a800")
LIGHT   = colors.HexColor("#f7f8fc")
BORDER  = colors.HexColor("#e2e6f0")
WHITE   = colors.white
GREY    = colors.HexColor("#6b7a99")

def _fmt_date(d) -> str:
    if not d:
        return "—"
    if isinstance(d, str):
        return d
    return d.strftime("%d-%b-%Y")

def _fmt_currency(v) -> str:
    if v is None:
        return "—"
    try:
        return f"₹{float(v):,.2f}"
    except Exception:
        return str(v)

def _fmt_num(v) -> str:
    if v is None:
        return "—"
    try:
        return f"{int(v):,}"
    except Exception:
        return str(v)

def _safe(v, default="—") -> str:
    return str(v) if v not in (None, "", []) else default


# ══════════════════════════════════════════════════════════════════════════════
# PDF GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def _build_pdf_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RegisterTitle",
        fontName="Helvetica-Bold",
        fontSize=14,
        textColor=NAVY,
        spaceAfter=4,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "RegisterSub",
        fontName="Helvetica",
        fontSize=9,
        textColor=GREY,
        spaceAfter=2,
        alignment=TA_CENTER,
    )
    section_style = ParagraphStyle(
        "SectionHead",
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=NAVY,
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "Body",
        fontName="Helvetica",
        fontSize=8.5,
        textColor=colors.black,
        leading=12,
    )
    footer_style = ParagraphStyle(
        "Footer",
        fontName="Helvetica-Oblique",
        fontSize=7.5,
        textColor=GREY,
        alignment=TA_CENTER,
    )
    return title_style, sub_style, section_style, body_style, footer_style


def _header_table(company: Dict, register_name: str, register_ref: str):
    """Top header block for each register PDF."""
    title_s, sub_s, _, _, _ = _build_pdf_styles()

    header_data = [[
        Paragraph(f"<b>{register_name}</b>", title_s),
    ]]
    sub_data = [[
        Paragraph(f"{company.get('company_name', '')} | CIN: {company.get('cin') or 'To be updated'}", sub_s),
        Paragraph(f"{register_ref} | As on {_fmt_date(date.today())}", sub_s),
    ]]

    t = Table([[Paragraph(f"<b>{register_name}</b>", title_s)]], colWidths=[None])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    sub_title_s = ParagraphStyle("sub2", fontName="Helvetica", fontSize=9, textColor=GREY, alignment=TA_CENTER)
    sub_t = Table([[
        Paragraph(f"{company.get('company_name', '')} &nbsp;|&nbsp; CIN: {company.get('cin') or 'N/A'} &nbsp;|&nbsp; {register_ref} &nbsp;|&nbsp; As on {_fmt_date(date.today())}", sub_title_s),
    ]])
    sub_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    return [t, sub_t, Spacer(1, 8)]


def _styled_table(data: List[List], col_widths: List, header_rows: int = 1):
    """Create a ReportLab table with standard CA firm styling."""
    t = Table(data, colWidths=col_widths, repeatRows=header_rows)
    style = [
        # Header row
        ("BACKGROUND",  (0, 0), (-1, header_rows - 1), NAVY),
        ("TEXTCOLOR",   (0, 0), (-1, header_rows - 1), WHITE),
        ("FONTNAME",    (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, header_rows - 1), 8),
        ("ALIGN",       (0, 0), (-1, header_rows - 1), "CENTER"),
        ("TOPPADDING",  (0, 0), (-1, header_rows - 1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, header_rows - 1), 7),
        # Data rows
        ("FONTNAME",    (0, header_rows), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, header_rows), (-1, -1), 8),
        ("ALIGN",       (0, header_rows), (-1, -1), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, header_rows), (-1, -1), 5),
        ("BOTTOMPADDING", (0, header_rows), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [WHITE, LIGHT]),
        # Grid
        ("GRID",        (0, 0), (-1, -1), 0.4, BORDER),
        ("LINEBELOW",   (0, header_rows - 1), (-1, header_rows - 1), 1.2, NAVY),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    t.setStyle(TableStyle(style))
    return t


def _missing_info_notice(fields: List[str]) -> List:
    """Return a warning block if data is missing."""
    if not fields:
        return []
    _, _, _, _, footer_s = _build_pdf_styles()
    warn = ParagraphStyle("warn", fontName="Helvetica-Oblique", fontSize=8, textColor=colors.HexColor("#c2410c"))
    return [
        Spacer(1, 6),
        Paragraph("⚠ The following information is missing and should be updated in Company Master:", warn),
        Paragraph(" • " + "  •  ".join(fields), warn),
        Spacer(1, 4),
    ]


# ──────────────────────────────────────────────────────────────────────────────
# 1. REGISTER OF MEMBERS (MGT-1 / Section 88)
# ──────────────────────────────────────────────────────────────────────────────

def generate_members_pdf(company: Dict, shareholders: List[Dict]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    _, _, section_s, body_s, footer_s = _build_pdf_styles()
    story = _header_table(company, "Register of Members", "MGT-1 | Section 88, Companies Act 2013")

    missing = []
    if not shareholders:
        missing.append("Shareholder records")

    if missing:
        story += _missing_info_notice(missing)

    # Table
    headers = ["Folio No.", "Name of Member", "Address", "PAN", "Type",
               "Class of Shares", "No. of Shares", "Shareholding %",
               "Date of Allotment", "Distinctive Nos."]
    col_w   = [22*mm, 40*mm, 50*mm, 22*mm, 20*mm, 22*mm, 22*mm, 22*mm, 25*mm, 30*mm]

    rows = [headers]
    total_shares = 0
    for i, s in enumerate(shareholders, 1):
        total_shares += int(s.get("number_of_shares") or 0)
        rows.append([
            _safe(s.get("folio_number")),
            _safe(s.get("full_name")),
            _safe(s.get("address") or s.get("city") or ""),
            _safe(s.get("pan")),
            _safe(s.get("shareholder_type")),
            _safe(s.get("class_of_shares", "Equity")),
            _fmt_num(s.get("number_of_shares")),
            f"{float(s.get('shareholding_ratio') or 0):.4f}%",
            _fmt_date(s.get("date_of_allotment")),
            "—",  # Distinctive numbers — to be filled manually
        ])

    # Totals row
    rows.append(["", "TOTAL", "", "", "", "", _fmt_num(total_shares), "100.0000%", "", ""])
    t = _styled_table(rows, col_w)
    # Style totals row
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, len(rows)-1), (-1, len(rows)-1), LIGHT),
        ("FONTNAME",   (0, len(rows)-1), (-1, len(rows)-1), "Helvetica-Bold"),
        ("LINEABOVE",  (0, len(rows)-1), (-1, len(rows)-1), 1, NAVY),
    ]))
    story.append(t)

    story += [
        Spacer(1, 12),
        Paragraph("Authenticated by: ___________________________  Date: _______________  Designation: _______________", footer_s),
        Paragraph("This Register is maintained pursuant to Section 88 of the Companies Act, 2013 and Rule 3 of the Companies (Management and Administration) Rules, 2014.", footer_s),
    ]

    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# 2. REGISTER OF DIRECTORS & KMP (Section 170 / Rule 17)
# ──────────────────────────────────────────────────────────────────────────────

def generate_directors_pdf(company: Dict, directors: List[Dict]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    _, _, section_s, body_s, footer_s = _build_pdf_styles()
    story = _header_table(company, "Register of Directors and Key Managerial Personnel",
                          "MBP-1 / Section 170, Companies Act 2013")

    headers = ["Sr.", "Name of Director", "DIN", "Designation", "PAN",
               "Date of Appointment", "Date of Cessation",
               "Address", "Nationality", "Shares Held", "DSC Expiry"]
    col_w   = [8*mm, 38*mm, 20*mm, 30*mm, 20*mm, 26*mm, 26*mm, 45*mm, 18*mm, 18*mm, 20*mm]

    rows = [headers]
    for i, d in enumerate(directors, 1):
        rows.append([
            str(i),
            _safe(d.get("full_name")),
            _safe(d.get("din")),
            _safe(d.get("designation", "Director")),
            _safe(d.get("pan")),
            _fmt_date(d.get("date_of_appointment")),
            _fmt_date(d.get("date_of_cessation")) if d.get("date_of_cessation") else "—",
            _safe(d.get("residential_address") or d.get("city") or ""),
            _safe(d.get("nationality", "Indian")),
            _fmt_num(d.get("number_of_shares")),
            _fmt_date(d.get("dsc_expiry_date")),
        ])

    story.append(_styled_table(rows, col_w))
    story += [
        Spacer(1, 12),
        Paragraph("Authenticated by: ___________________________  Date: _______________  Designation: _______________", footer_s),
        Paragraph("This Register is maintained pursuant to Section 170 of the Companies Act, 2013 and Rule 17 of the Companies (Appointment and Qualification of Directors) Rules, 2014.", footer_s),
    ]
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# 3. REGISTER OF CHARGES (CHG-7 / Section 85)
# ──────────────────────────────────────────────────────────────────────────────

def generate_charges_pdf(company: Dict, charges: List[Dict]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    _, _, section_s, body_s, footer_s = _build_pdf_styles()
    story = _header_table(company, "Register of Charges",
                          "CHG-7 | Section 85, Companies Act 2013")

    headers = ["Sr.", "MCA Charge ID", "Charge Holder (Lender)",
               "Amount (₹)", "Date of Creation",
               "Property / Assets Charged",
               "Date of Satisfaction", "Status", "Remarks"]
    col_w   = [8*mm, 22*mm, 40*mm, 25*mm, 25*mm, 65*mm, 25*mm, 18*mm, 25*mm]

    rows = [headers]
    total_active = 0
    for i, c in enumerate(charges, 1):
        if str(c.get("status", "")).lower() == "active":
            total_active += float(c.get("charge_amount") or 0)
        rows.append([
            str(i),
            _safe(c.get("charge_id")),
            _safe(c.get("charge_holder")),
            _fmt_currency(c.get("charge_amount")),
            _fmt_date(c.get("date_of_creation")),
            _safe(c.get("property_charged")),
            _fmt_date(c.get("date_of_satisfaction")),
            _safe(c.get("status", "Active")),
            _safe(c.get("remarks")),
        ])

    story.append(_styled_table(rows, col_w))
    story += [
        Spacer(1, 6),
        Paragraph(f"Total Active Charge Amount: {_fmt_currency(total_active)}", footer_s),
        Spacer(1, 10),
        Paragraph("Authenticated by: ___________________________  Date: _______________  Designation: _______________", footer_s),
        Paragraph("This Register is maintained pursuant to Section 85 of the Companies Act, 2013.", footer_s),
    ]
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# 4. REGISTER OF SHARE TRANSFERS (SH-6 / Section 56)
# ──────────────────────────────────────────────────────────────────────────────

def generate_transfers_pdf(company: Dict, transfers: List[Dict]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    _, _, section_s, body_s, footer_s = _build_pdf_styles()
    story = _header_table(company, "Register of Share Transfers",
                          "SH-6 | Section 56, Companies Act 2013")

    headers = ["Sr.", "Transfer Deed No.", "Date of Transfer",
               "Transferor (From)", "Transferee (To)",
               "No. of Shares", "Price/Share (₹)",
               "Total Consideration (₹)", "Remarks"]
    col_w   = [8*mm, 25*mm, 25*mm, 40*mm, 40*mm, 22*mm, 22*mm, 28*mm, 30*mm]

    rows = [headers]
    total_shares = 0
    total_consideration = 0
    for i, t in enumerate(transfers, 1):
        shares = int(t.get("number_of_shares") or 0)
        price  = float(t.get("transfer_price_per_share") or 0)
        consideration = float(t.get("consideration_amount") or 0) or (shares * price)
        total_shares += shares
        total_consideration += consideration
        rows.append([
            str(i),
            _safe(t.get("transfer_deed_number")),
            _fmt_date(t.get("transfer_date")),
            _safe(t.get("from_name")),
            _safe(t.get("to_name")),
            _fmt_num(shares),
            _fmt_currency(price) if price else "—",
            _fmt_currency(consideration) if consideration else "—",
            _safe(t.get("remarks")),
        ])

    rows.append(["", "TOTALS", "", "", "", _fmt_num(total_shares), "", _fmt_currency(total_consideration), ""])
    t = _styled_table(rows, col_w)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, len(rows)-1), (-1, len(rows)-1), LIGHT),
        ("FONTNAME",   (0, len(rows)-1), (-1, len(rows)-1), "Helvetica-Bold"),
        ("LINEABOVE",  (0, len(rows)-1), (-1, len(rows)-1), 1, NAVY),
    ]))
    story.append(t)
    story += [
        Spacer(1, 10),
        Paragraph("Authenticated by: ___________________________  Date: _______________  Designation: _______________", footer_s),
        Paragraph("This Register is maintained pursuant to Section 56 of the Companies Act, 2013 and Rule 11 of the Companies (Share Capital and Debentures) Rules, 2014.", footer_s),
    ]
    doc.build(story)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

NAVY_HEX   = "0F1F3D"
GOLD_HEX   = "E8A800"
LIGHT_HEX  = "F7F8FC"
BORDER_HEX = "E2E6F0"
WHITE_HEX  = "FFFFFF"
RED_HEX    = "FEF2F2"


def _xl_border(color="E2E6F0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_header_font():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=10)


def _xl_header_fill():
    return PatternFill("solid", fgColor=NAVY_HEX)


def _xl_gold_fill():
    return PatternFill("solid", fgColor=GOLD_HEX)


def _xl_light_fill():
    return PatternFill("solid", fgColor=LIGHT_HEX)


def _xl_title_row(ws, title: str, subtitle: str, cols: int):
    """Write title and subtitle rows with merged cells."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY_HEX)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=9, color="6B7A99", italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_HEX)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)  # spacer
    ws.row_dimensions[3].height = 6


def _xl_write_headers(ws, headers: List[str], row: int = 4, col_widths: List[int] = None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = _xl_header_font()
        cell.fill = _xl_header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _xl_border("344D6E")
        if col_widths and ci <= len(col_widths):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.row_dimensions[row].height = 28


def _xl_write_row(ws, row_idx: int, values: List, alt: bool = False):
    fill = _xl_light_fill() if alt else PatternFill("solid", fgColor=WHITE_HEX)
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _xl_border()
    ws.row_dimensions[row_idx].height = 18


def _xl_total_row(ws, row_idx: int, values: List, cols: int):
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        cell.font = Font(name="Calibri", bold=True, size=9, color=NAVY_HEX)
        cell.fill = _xl_gold_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _xl_border(GOLD_HEX)
    ws.row_dimensions[row_idx].height = 20


# ── Register of Members — Excel ───────────────────────────────────────────────

def generate_members_excel(company: Dict, shareholders: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Register of Members"
    ws.sheet_view.showGridLines = False

    cols = 10
    _xl_title_row(ws,
        f"Register of Members — {company.get('company_name', '')}",
        f"MGT-1 | CIN: {company.get('cin') or 'N/A'} | Section 88, Companies Act 2013 | As on {_fmt_date(date.today())}",
        cols)

    headers = ["Folio No.", "Name of Member", "Address", "PAN",
               "Type", "Class", "No. of Shares", "Shareholding %",
               "Date of Allotment", "Distinctive Nos."]
    widths  = [14, 30, 35, 14, 16, 16, 16, 14, 18, 18]
    _xl_write_headers(ws, headers, row=4, col_widths=widths)

    total_shares = 0
    for i, s in enumerate(shareholders):
        total_shares += int(s.get("number_of_shares") or 0)
        _xl_write_row(ws, 5 + i, [
            _safe(s.get("folio_number")),
            _safe(s.get("full_name")),
            _safe(s.get("address") or s.get("city") or ""),
            _safe(s.get("pan")),
            _safe(s.get("shareholder_type")),
            _safe(s.get("class_of_shares", "Equity")),
            int(s.get("number_of_shares") or 0),
            f"{float(s.get('shareholding_ratio') or 0):.4f}%",
            _fmt_date(s.get("date_of_allotment")),
            "—",
        ], alt=(i % 2 == 1))

    total_row = 5 + len(shareholders)
    _xl_total_row(ws, total_row, ["", "TOTAL", "", "", "", "", total_shares, "100.0000%", "", ""], cols)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(cols)}{total_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Register of Directors — Excel ─────────────────────────────────────────────

def generate_directors_excel(company: Dict, directors: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Register of Directors"
    ws.sheet_view.showGridLines = False

    cols = 11
    _xl_title_row(ws,
        f"Register of Directors & KMP — {company.get('company_name', '')}",
        f"Section 170 | CIN: {company.get('cin') or 'N/A'} | Companies Act 2013 | As on {_fmt_date(date.today())}",
        cols)

    headers = ["Sr.", "Name", "DIN", "Designation", "PAN",
               "Date of Appointment", "Date of Cessation",
               "Address", "Nationality", "Shares Held", "DSC Expiry"]
    widths  = [6, 28, 14, 22, 14, 18, 18, 35, 14, 14, 16]
    _xl_write_headers(ws, headers, row=4, col_widths=widths)

    for i, d in enumerate(directors):
        _xl_write_row(ws, 5 + i, [
            i + 1,
            _safe(d.get("full_name")),
            _safe(d.get("din")),
            _safe(d.get("designation", "Director")),
            _safe(d.get("pan")),
            _fmt_date(d.get("date_of_appointment")),
            _fmt_date(d.get("date_of_cessation")) if d.get("date_of_cessation") else "—",
            _safe(d.get("residential_address") or d.get("city") or ""),
            _safe(d.get("nationality", "Indian")),
            int(d.get("number_of_shares") or 0),
            _fmt_date(d.get("dsc_expiry_date")),
        ], alt=(i % 2 == 1))

    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Register of Charges — Excel ───────────────────────────────────────────────

def generate_charges_excel(company: Dict, charges: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Register of Charges"
    ws.sheet_view.showGridLines = False

    cols = 9
    _xl_title_row(ws,
        f"Register of Charges — {company.get('company_name', '')}",
        f"CHG-7 | CIN: {company.get('cin') or 'N/A'} | Section 85, Companies Act 2013 | As on {_fmt_date(date.today())}",
        cols)

    headers = ["Sr.", "MCA Charge ID", "Charge Holder",
               "Amount (₹)", "Date of Creation",
               "Property Charged", "Date of Satisfaction", "Status", "Remarks"]
    widths  = [6, 18, 30, 18, 18, 40, 18, 14, 24]
    _xl_write_headers(ws, headers, row=4, col_widths=widths)

    total_active = 0
    for i, c in enumerate(charges):
        amt = float(c.get("charge_amount") or 0)
        if str(c.get("status", "")).lower() == "active":
            total_active += amt

        status_val = _safe(c.get("status", "Active"))
        row_idx = 5 + i
        _xl_write_row(ws, row_idx, [
            i + 1,
            _safe(c.get("charge_id")),
            _safe(c.get("charge_holder")),
            amt if amt else "—",
            _fmt_date(c.get("date_of_creation")),
            _safe(c.get("property_charged")),
            _fmt_date(c.get("date_of_satisfaction")),
            status_val,
            _safe(c.get("remarks")),
        ], alt=(i % 2 == 1))

        # Highlight satisfied charges in green
        if status_val == "Satisfied":
            for ci in range(1, cols + 1):
                ws.cell(row=row_idx, column=ci).fill = PatternFill("solid", fgColor="F0FDF4")
        elif status_val == "Active":
            ws.cell(row=row_idx, column=8).font = Font(name="Calibri", bold=True, size=9, color="DC2626")

    total_row = 5 + len(charges)
    _xl_total_row(ws, total_row,
        ["", "TOTAL ACTIVE CHARGES", "", f"₹{total_active:,.2f}", "", "", "", "", ""], cols)

    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Register of Share Transfers — Excel ───────────────────────────────────────

def generate_transfers_excel(company: Dict, transfers: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Share Transfer Register"
    ws.sheet_view.showGridLines = False

    cols = 9
    _xl_title_row(ws,
        f"Register of Share Transfers — {company.get('company_name', '')}",
        f"SH-6 | CIN: {company.get('cin') or 'N/A'} | Section 56, Companies Act 2013 | As on {_fmt_date(date.today())}",
        cols)

    headers = ["Sr.", "Deed No.", "Date of Transfer",
               "Transferor (From)", "Transferee (To)",
               "No. of Shares", "Price/Share (₹)",
               "Total Consideration (₹)", "Remarks"]
    widths  = [6, 18, 18, 30, 30, 16, 16, 22, 24]
    _xl_write_headers(ws, headers, row=4, col_widths=widths)

    total_shares = 0
    total_consideration = 0
    for i, t in enumerate(transfers):
        shares = int(t.get("number_of_shares") or 0)
        price  = float(t.get("transfer_price_per_share") or 0)
        consid = float(t.get("consideration_amount") or 0) or (shares * price)
        total_shares += shares
        total_consideration += consid
        _xl_write_row(ws, 5 + i, [
            i + 1,
            _safe(t.get("transfer_deed_number")),
            _fmt_date(t.get("transfer_date")),
            _safe(t.get("from_name")),
            _safe(t.get("to_name")),
            shares,
            price if price else "—",
            consid if consid else "—",
            _safe(t.get("remarks")),
        ], alt=(i % 2 == 1))

    total_row = 5 + len(transfers)
    _xl_total_row(ws, total_row,
        ["", "TOTALS", "", "", "", total_shares, "", f"₹{total_consideration:,.2f}", ""], cols)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(cols)}{total_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# COMBINED EXCEL WORKBOOK (all 4 registers in one file)
# ══════════════════════════════════════════════════════════════════════════════

def generate_all_registers_excel(
    company: Dict,
    shareholders: List[Dict],
    directors: List[Dict],
    charges: List[Dict],
    transfers: List[Dict],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Cover Sheet ──────────────────────────────────────────────────────────
    cover = wb.create_sheet("Cover")
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 60

    cover.merge_cells("A1:A3")
    cover["A1"] = "STATUTORY REGISTERS"
    cover["A1"].font = Font(name="Calibri", bold=True, size=22, color="FFFFFF")
    cover["A1"].fill = PatternFill("solid", fgColor=NAVY_HEX)
    cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    cover.row_dimensions[1].height = 60

    cover["A4"] = company.get("company_name", "")
    cover["A4"].font = Font(name="Calibri", bold=True, size=14)
    cover["A4"].alignment = Alignment(horizontal="center")
    cover.row_dimensions[4].height = 24

    cover["A5"] = f"CIN: {company.get('cin') or 'N/A'}"
    cover["A5"].font = Font(name="Calibri", size=11, color="6B7A99")
    cover["A5"].alignment = Alignment(horizontal="center")

    cover["A6"] = f"Registered Office: {company.get('registered_office_address') or 'N/A'}"
    cover["A6"].font = Font(name="Calibri", size=10, color="6B7A99")
    cover["A6"].alignment = Alignment(horizontal="center")

    cover["A8"] = f"Generated on: {_fmt_date(date.today())}"
    cover["A8"].font = Font(name="Calibri", italic=True, size=10, color="6B7A99")
    cover["A8"].alignment = Alignment(horizontal="center")

    registers = [
        ("1. Register of Members",          "MGT-1 | Section 88"),
        ("2. Register of Directors & KMP",  "Section 170"),
        ("3. Register of Charges",          "CHG-7 | Section 85"),
        ("4. Register of Share Transfers",  "SH-6 | Section 56"),
    ]
    for row_i, (name, ref) in enumerate(registers, 10):
        cover[f"A{row_i}"] = f"  {name} — {ref}"
        cover[f"A{row_i}"].font = Font(name="Calibri", size=10)
        cover.row_dimensions[row_i].height = 18

    # ── Individual sheets ────────────────────────────────────────────────────
    def _add_members(wb):
        ws = wb.create_sheet("Register of Members")
        ws.sheet_view.showGridLines = False
        cols = 10
        _xl_title_row(ws, f"Register of Members — {company.get('company_name','')}",
            f"MGT-1 | CIN: {company.get('cin') or 'N/A'} | Section 88 | As on {_fmt_date(date.today())}", cols)
        headers = ["Folio No.","Name of Member","Address","PAN","Type","Class","No. of Shares","Holding %","Date of Allotment","Distinctive Nos."]
        _xl_write_headers(ws, headers, 4, [14,30,35,14,16,16,16,14,18,18])
        total_shares = 0
        for i,s in enumerate(shareholders):
            total_shares += int(s.get("number_of_shares") or 0)
            _xl_write_row(ws, 5+i, [_safe(s.get("folio_number")),_safe(s.get("full_name")),_safe(s.get("address") or s.get("city") or ""),_safe(s.get("pan")),_safe(s.get("shareholder_type")),_safe(s.get("class_of_shares","Equity")),int(s.get("number_of_shares") or 0),f"{float(s.get('shareholding_ratio') or 0):.4f}%",_fmt_date(s.get("date_of_allotment")),"—"], alt=(i%2==1))
        _xl_total_row(ws, 5+len(shareholders), ["","TOTAL","","","","",total_shares,"100.0000%","",""], cols)
        ws.freeze_panes = "A5"

    def _add_directors(wb):
        ws = wb.create_sheet("Register of Directors")
        ws.sheet_view.showGridLines = False
        cols = 11
        _xl_title_row(ws, f"Register of Directors & KMP — {company.get('company_name','')}",
            f"Section 170 | CIN: {company.get('cin') or 'N/A'} | As on {_fmt_date(date.today())}", cols)
        headers = ["Sr.","Name","DIN","Designation","PAN","Date of Appt.","Date of Cessation","Address","Nationality","Shares Held","DSC Expiry"]
        _xl_write_headers(ws, headers, 4, [6,28,14,22,14,18,18,35,14,14,16])
        for i,d in enumerate(directors):
            _xl_write_row(ws, 5+i, [i+1,_safe(d.get("full_name")),_safe(d.get("din")),_safe(d.get("designation","Director")),_safe(d.get("pan")),_fmt_date(d.get("date_of_appointment")),_fmt_date(d.get("date_of_cessation")) if d.get("date_of_cessation") else "—",_safe(d.get("residential_address") or d.get("city") or ""),_safe(d.get("nationality","Indian")),int(d.get("number_of_shares") or 0),_fmt_date(d.get("dsc_expiry_date"))], alt=(i%2==1))
        ws.freeze_panes = "A5"

    def _add_charges(wb):
        ws = wb.create_sheet("Register of Charges")
        ws.sheet_view.showGridLines = False
        cols = 9
        _xl_title_row(ws, f"Register of Charges — {company.get('company_name','')}",
            f"CHG-7 | CIN: {company.get('cin') or 'N/A'} | Section 85 | As on {_fmt_date(date.today())}", cols)
        headers = ["Sr.","MCA Charge ID","Charge Holder","Amount (₹)","Date of Creation","Property Charged","Date of Satisfaction","Status","Remarks"]
        _xl_write_headers(ws, headers, 4, [6,18,30,18,18,40,18,14,24])
        total_active = 0
        for i,c in enumerate(charges):
            amt = float(c.get("charge_amount") or 0)
            if str(c.get("status","")).lower() == "active": total_active += amt
            _xl_write_row(ws, 5+i, [i+1,_safe(c.get("charge_id")),_safe(c.get("charge_holder")),amt if amt else "—",_fmt_date(c.get("date_of_creation")),_safe(c.get("property_charged")),_fmt_date(c.get("date_of_satisfaction")),_safe(c.get("status","Active")),_safe(c.get("remarks"))], alt=(i%2==1))
        _xl_total_row(ws, 5+len(charges), ["","TOTAL ACTIVE","",f"₹{total_active:,.2f}","","","","",""], cols)
        ws.freeze_panes = "A5"

    def _add_transfers(wb):
        ws = wb.create_sheet("Share Transfer Register")
        ws.sheet_view.showGridLines = False
        cols = 9
        _xl_title_row(ws, f"Register of Share Transfers — {company.get('company_name','')}",
            f"SH-6 | CIN: {company.get('cin') or 'N/A'} | Section 56 | As on {_fmt_date(date.today())}", cols)
        headers = ["Sr.","Deed No.","Date of Transfer","Transferor","Transferee","No. of Shares","Price/Share","Consideration (₹)","Remarks"]
        _xl_write_headers(ws, headers, 4, [6,18,18,30,30,16,16,22,24])
        total_s, total_c = 0, 0
        for i,t in enumerate(transfers):
            s = int(t.get("number_of_shares") or 0); p = float(t.get("transfer_price_per_share") or 0)
            c = float(t.get("consideration_amount") or 0) or (s*p)
            total_s += s; total_c += c
            _xl_write_row(ws, 5+i, [i+1,_safe(t.get("transfer_deed_number")),_fmt_date(t.get("transfer_date")),_safe(t.get("from_name")),_safe(t.get("to_name")),s,p if p else "—",c if c else "—",_safe(t.get("remarks"))], alt=(i%2==1))
        _xl_total_row(ws, 5+len(transfers), ["","TOTALS","","","",total_s,"",f"₹{total_c:,.2f}",""], cols)
        ws.freeze_panes = "A5"

    _add_members(wb)
    _add_directors(wb)
    _add_charges(wb)
    _add_transfers(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
